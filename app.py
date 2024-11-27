from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
import os
import random

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
ALLOWED_EXTENSIONS = {'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['RESULT_FOLDER'] = RESULT_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

if not os.path.exists(RESULT_FOLDER):
    os.makedirs(RESULT_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_xlsx(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({'name': row[0], 'email': row[1]})
    return data

def generate_secret_santa(employees, previous_assignments):
    available = employees[:]
    assignments = {}

    for emp in employees:
        choices = [e for e in available if e['name'] != emp['name'] and e not in previous_assignments.get(emp['name'], [])]
        if not choices:
            raise ValueError("Cannot generate valid Secret Santa assignments.")
        chosen = random.choice(choices)
        assignments[emp['name']] = chosen
        available.remove(chosen)
    return assignments

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        employee_file = request.files.get('employee_file')
        previous_file = request.files.get('previous_file')

        if not (employee_file and allowed_file(employee_file.filename)):
            flash("Please upload a valid employee XLSX file.", "error")
            return redirect(request.url)
        if not (previous_file and allowed_file(previous_file.filename)):
            flash("Please upload a valid previous assignments XLSX file.", "error")
            return redirect(request.url)

        emp_filename = secure_filename(employee_file.filename)
        prev_filename = secure_filename(previous_file.filename)

        emp_file_path = os.path.join(app.config['UPLOAD_FOLDER'], emp_filename)
        prev_file_path = os.path.join(app.config['UPLOAD_FOLDER'], prev_filename)

        employee_file.save(emp_file_path)
        previous_file.save(prev_file_path)

        try:
            employees = parse_xlsx(emp_file_path)
            previous_data = parse_xlsx(prev_file_path)

            previous_assignments = {}
            for item in previous_data:
                giver = item['name']
                receiver = {'name': item['email'], 'email': item['email']}
                if giver not in previous_assignments:
                    previous_assignments[giver] = []
                previous_assignments[giver].append(receiver)

            assignments = generate_secret_santa(employees, previous_assignments)

            result_path = os.path.join(app.config['RESULT_FOLDER'], 'secret_santa_result.xlsx')
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Employee_Name', 'Employee_EmailID', 'Secret_Child_Name', 'Secret_Child_EmailID'])

            for giver, receiver in assignments.items():
                sheet.append([giver, receiver['email'], receiver['name'], receiver['email']])

            workbook.save(result_path)
            flash("Secret Santa assignments generated successfully! You can download File using above Link", "success")
            return render_template('result.html', result_file='secret_santa_result.xlsx')

        except Exception as e:
            flash(str(e), "error")
            return redirect(url_for('error', message=str(e)))

    return render_template('index.html')

@app.route('/download/<filename>')
def download(filename):
    file_path = os.path.join(app.config['RESULT_FOLDER'], filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    flash("The requested file does not exist.", "error")
    return redirect(url_for('index'))


if __name__ == '__main__':
    app.run(debug=True)
